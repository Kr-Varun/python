# -*- coding: utf-8 -*-
"""
Created on Mon Jun 11 03:37:16 2018

@author: unmanaged 24
"""
from openpyxl import Workbook
from openpyxl import load_workbook

from openpyxl.styles import Color, PatternFill, Font, Border,Fill 
from openpyxl.styles.colors import RED
import calendar
import random
import datetime

path="C:\\Users\\unmanaged 24\\Desktop\\python_code\\logs\\reports\\AllstateAnnualReports.xlsx"


wb=Workbook()  #creation of object for workbook class

for month in calendar.month_name:
    print(month)
    if not(len(str(month))==0):
        wb.create_sheet(month+"_2018")
    
wb.save(path)
wb.close()

wbRef=load_workbook(path,read_only=False)
for sheet in wbRef.get_sheet_names():
    print(sheet)
    
currentMonth=datetime.date.today().strftime("%B")+ "_2018"
sheetRef=wbRef.get_sheet_by_name(currentMonth)

for row in range(1,100):
    for col in range(1,10):
        sheetRef.cell(column=col,row=row,value="%d" %(random.randint(100,10000)))
wbRef.save(path)

for row in range(1,100):
    for col in range(1,10):
        cell=sheetRef.cell(row=row, column=col)
        if(int(cell.value)>5000):
            cell.font=Font(size=18,color=RED)
wbRef.save(path)
wbRef.close()





